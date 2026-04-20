from copy import copy
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "generated_certs"

TEMPLATE_FILES = {
    "Brandon Regular": "Cert Template Revised (Brandon).xlsx",
    "Hugo Regular": "Cert Template Revised (Hugo).xlsx",
    "North Regular": "Cert Template Revised (North).xlsx",
}

TEMPLATE_DEFAULTS = {
    "Brandon Regular": {
        "technician": "MT-1",
        "standard_1": "FLUKE/C19623",
        "standard_2": "REED/241209830/11-08-26",
        "standard_3": "TRACEABLE/250534900/09-02-27",
    },
    "Hugo Regular": {
        "technician": "MT-2",
        "standard_1": "FLUKE/A8B413",
        "standard_2": "REED/210205965/11-08-26",
        "standard_3": "TRACEABLE/250534899/09-02-27",
    },
    "North Regular": {
        "technician": "MT-1",
        "standard_1": "FLUKE/9103/C19635",
        "standard_2": "REED/C-370/241209386",
        "standard_3": "TRACEABLE/250534896/09-02-27",
    },
}

FIELD_MAP = {
    "company": "F4",
    "address": "F5",
    "city_state_zip": "F6",
    "technician": "F7",
    "manufacturer": "E9",
    "instrument": "E10",
    "model_number": "E11",
    "size_range": "E12",
    "serial_number": "E13",
    "identification": "E14",
    "standard_1": "E16",
    "standard_2": "E17",
    "standard_3": "E18",
}

# These are the exact cells from your template screenshot block.
OPTIONAL_MAP = {
    "date_calibrated": "K4",
    "calibration_interval": "K5",
    "customer_req_due": "K6",
    "invoice_number": "K7",
    "procedure": "K9",
    "rated_tolerance": "K10",
    "tolerance_as_found": "K11",
    "adjustments_made": "K12",
    "condition_as_found": "K13",
    "location": "K14",
    "temperature": "K15",
    "relative_humidity": "K16",
    "certificate_number": "K17",
    "certificate_issue_date": "K18",
}

REQUIRED_ALWAYS_FILLED_DEFAULTS = {
    "procedure": "MCP-1",
    "rated_tolerance": "±1°F",
    "tolerance_as_found": "IN",
    "adjustments_made": "NO",
    "condition_as_found": "FAIR",
    "location": "ON-SITE",
    "temperature": "N/A",
    "relative_humidity": "N/A",
}

FIRST_RESULT_ROW = 23
TEMPLATE_RESULT_ROWS = 3


def get_template_path(template_name: str) -> Path:
    if template_name not in TEMPLATE_FILES:
        valid = ", ".join(TEMPLATE_FILES)
        raise ValueError(f"Unknown template '{template_name}'. Choose one of: {valid}")
    return BASE_DIR / TEMPLATE_FILES[template_name]


def get_template_defaults(template_name: str) -> dict:
    return dict(TEMPLATE_DEFAULTS.get(template_name, {}))


def to_caps(value):
    if value is None:
        return ""
    return str(value).upper()


def normalize_date_string(value):
    if not value:
        return datetime.today().strftime("%m/%d/%Y")
    return str(value)


def apply_required_defaults(data: dict) -> dict:
    merged = dict(data)

    for key, default_value in REQUIRED_ALWAYS_FILLED_DEFAULTS.items():
        if not merged.get(key):
            merged[key] = default_value

    if not merged.get("certificate_issue_date"):
        merged["certificate_issue_date"] = datetime.today().strftime("%m/%d/%Y")

    return merged


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.number_format:
            dst.number_format = src.number_format
        if src.protection:
            dst.protection = copy(src.protection)


def calculate_error(actual_standard, uut):
    try:
        return round(float(uut) - float(actual_standard), 2)
    except (TypeError, ValueError):
        return None


def insert_result_rows(ws, results):
    extra_needed = max(0, len(results) - TEMPLATE_RESULT_ROWS)

    if extra_needed:
        insert_at = FIRST_RESULT_ROW + TEMPLATE_RESULT_ROWS
        ws.insert_rows(insert_at, amount=extra_needed)
        source_row = FIRST_RESULT_ROW + TEMPLATE_RESULT_ROWS - 1
        for i in range(extra_needed):
            copy_row_style(ws, source_row, insert_at + i)

    for offset, result in enumerate(results):
        row = FIRST_RESULT_ROW + offset
        point_no = result.get("point_no", offset + 1)
        actual_standard = result.get("actual_standard")
        as_found = result.get("as_found")
        error = result.get("error")
        uncertainty = result.get("uncertainty")

        if error is None:
            error = calculate_error(actual_standard, as_found)

        ws[f"C{row}"] = point_no
        ws[f"D{row}"] = actual_standard
        ws[f"H{row}"] = as_found
        ws[f"J{row}"] = error
        if uncertainty is not None:
            ws[f"K{row}"] = uncertainty


def fill_certificate(data: dict) -> Path:
    template_name = data.get("template_name", "Brandon Regular")
    template_path = get_template_path(template_name)

    if not template_path.exists():
        raise FileNotFoundError(
            f"Template not found: {template_path.name}. Put that template in the same folder as this file."
        )

    merged_data = get_template_defaults(template_name)
    merged_data.update({k: v for k, v in data.items() if v not in (None, "")})
    merged_data = apply_required_defaults(merged_data)

    wb = load_workbook(template_path)
    ws = wb["Table 1"]

    # Main/customer/instrument fields in ALL CAPS
    for key, cell in FIELD_MAP.items():
        if merged_data.get(key) is not None:
            ws[cell] = to_caps(merged_data[key])

    # Required template block in ALL CAPS and always filled
    for key, cell in OPTIONAL_MAP.items():
        if key == "certificate_issue_date":
            ws[cell] = to_caps(normalize_date_string(merged_data.get(key)))
        elif merged_data.get(key) is not None:
            ws[cell] = to_caps(merged_data[key])

    results = merged_data.get("results", [])
    if results:
        insert_result_rows(ws, results)

    OUTPUT_DIR.mkdir(exist_ok=True)

    serial = str(merged_data.get("serial_number", "UNKNOWN")).replace("/", "-")
    cert_number = str(merged_data.get("certificate_number", "")).strip().replace("/", "-")
    template_slug = template_name.lower().replace(" ", "_")
    parts = ["certificate", template_slug]
    if cert_number:
        parts.append(cert_number)
    parts.append(serial)

    out_path = OUTPUT_DIR / ("_".join(parts) + ".xlsx")
    wb.save(out_path)
    return out_path


def generate_certificates_for_job(job_data: dict, instruments: list[dict], template_name: str) -> list[Path]:
    created = []
    for instrument in instruments:
        merged = dict(job_data)
        merged.update(instrument)
        merged["template_name"] = template_name
        created.append(fill_certificate(merged))
    return created
